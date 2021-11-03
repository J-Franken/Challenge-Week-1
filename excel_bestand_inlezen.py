from openpyxl import load_workbook
import gegeven_antwoord_controleren

def excel_inlezen(bestandsnaam):
    sheet = (load_workbook(filename=bestandsnaam)).active
    LETTERS = ['A) ', 'B) ', 'C) ', 'D) ']
    BIJBEHORENDE_PUNTEN = {'A': 1,
                           'B': 2,
                           'C': 3,
                           'D': 4
                           }

    huidige_punten_dict = {'SE': 0,
                           'IAT': 0,
                           'FICT': 0,
                           'BDaM': 0
                           }
    e = 0
    # loopt door elke vraag heen
    # @ToDo vervang 92 met aantal rijen
    for vraag in range(1, 92, 6):
        print('\n')
        print(sheet['A' + str(vraag)].value)
        print('\n')
        e = 0
        # tijdens 1 vraag loop, loopt door alle antwoorden heen
        for i in range((vraag + 1), (vraag + 5)):
            # letter[e] zorgt ervoor dat de bijbehorende antwoorden de juist meerkeuzenotaties krijgen
            print(LETTERS[e] + (sheet["A" + str(i)].value))
            e = e + 1
        # antwoord vraagt de input van de gebruiker
        antwoord = gegeven_antwoord_controleren.antwoord_invoer_en_controle()

        specialisatie_key = sheet["B" + str((BIJBEHORENDE_PUNTEN[antwoord]) + vraag)].value
        punten_value = sheet["C" + str((BIJBEHORENDE_PUNTEN[antwoord]) + vraag)].value

        huidige_punten_dict[specialisatie_key] = huidige_punten_dict[specialisatie_key] + punten_value

    return huidige_punten_dict

def excel_vragen_en_antwoorden(bestandsnaam):
    #deze functie zorgt ervoor dat de vragen in de excel bestand opgeslagen worden
    vragen_lijst= {}
    sheet = (load_workbook(filename=bestandsnaam)).active
    for vraag in range(1, 92, 6):
        sheet['A' + str(vraag)].value
        










